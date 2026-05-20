"""
Analyze Rainforest-specific climate control scope.

This script counts:
1. Input parameters for Rainforest only
2. Output sensors for Rainforest only
3. Provides exact filter instructions for Excel
"""

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
WORKBOOK_PATH = ROOT / "Sensors_Description" / "variables_schema.xlsx"

def analyze_input_sheet():
    """Analyze Input sheet for Rainforest parameters."""
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["Input"]
    
    rainforest_controls = []
    categories = {
        "temperature": [],
        "fan": [],
        "valve_command": []
    }
    
    # Skip header rows (1-3), start from row 4
    for row_idx in range(4, ws.max_row + 1):
        physical_quantity = ws.cell(row_idx, 1).value
        role = ws.cell(row_idx, 7).value  # Column G - role

        # Check if it's Rainforest via role column (Column G)
        if role and "Rainforest" in str(role):
            # Check if it's a climate control parameter
            phys_lower = (physical_quantity or "").lower()
            
            if any(x in phys_lower for x in [
                "temperature setpoint",
                "supply fan command",
                "cooling valve command",
                "heating valve command",
                "economizer / damper control",
                "valve command / position",
                "occupancy / schedule command"
            ]):
                row_data = {
                    "row": row_idx,
                    "physical_quantity": physical_quantity,
                    "sensor_code": ws.cell(row_idx, 13).value,
                }
                rainforest_controls.append(row_data)
                
                # Categorize
                if "temperature" in phys_lower:
                    categories["temperature"].append(row_data)
                elif "fan" in phys_lower:
                    categories["fan"].append(row_data)
                else:
                    categories["valve_command"].append(row_data)
    
    return rainforest_controls, categories

def analyze_output_sheet():
    """Analyze Output sheet for Rainforest sensors."""
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["Output"]
    
    rainforest_sensors = []
    categories = {
        "temperature": [],
        "humidity": []
    }
    
    # Skip header rows (1-3), start from row 4
    for row_idx in range(4, ws.max_row + 1):
        physical_quantity = ws.cell(row_idx, 1).value
        sensor_code = ws.cell(row_idx, 13).value
        role = ws.cell(row_idx, 7).value  # Column G - role

        # All Output sensors are for RainForest (assumes Output sheet is RainForest-only)
        # But we can verify by checking if sensor_code starts with TRF
        if sensor_code and "TRF" in str(sensor_code):
            row_data = {
                "row": row_idx,
                "physical_quantity": physical_quantity,
                "sensor_code": sensor_code,
            }
            rainforest_sensors.append(row_data)
            
            # Categorize
            phys_lower = (physical_quantity or "").lower()
            if "temperature" in phys_lower or "temp" in phys_lower:
                categories["temperature"].append(row_data)
            elif "humidity" in phys_lower or "rh" in phys_lower:
                categories["humidity"].append(row_data)
    
    return rainforest_sensors, categories

def main():
    print("=" * 80)
    print("RAINFOREST CLIMATE CONTROL PROJECT SCOPE")
    print("=" * 80)
    
    # Analyze Input sheet
    print("\n[INPUT SHEET - CONTROL PARAMETERS]")
    input_controls, input_cats = analyze_input_sheet()
    print(f"Total Input parameters for Rainforest: {len(input_controls)}")
    print(f"  🌡️  Temperature controls: {len(input_cats['temperature'])}")
    print(f"  🌬️  Fan controls: {len(input_cats['fan'])}")
    print(f"  🔧 Valve/command controls: {len(input_cats['valve_command'])}")
    
    # Analyze Output sheet
    print("\n[OUTPUT SHEET - MONITORING SENSORS]")
    output_sensors, output_cats = analyze_output_sheet()
    print(f"Total Output sensors for Rainforest: {len(output_sensors)}")
    print(f"  🌡️  Temperature sensors: {len(output_cats['temperature'])}")
    print(f"  💧 Humidity sensors: {len(output_cats['humidity'])}")
    
    # Summary
    print("\n" + "=" * 80)
    print("RAINFOREST CLIMATE CONTROL GOAL:")
    print("=" * 80)
    print(f"\nInput control parameters:  {len(input_controls)}")
    print(f"           ↓")
    print(f"   RainForest Climate System")
    print(f"           ↓")
    print(f"Output monitoring sensors: {len(output_sensors)}")
    print(f"\nFormula:")
    print(f"Input ({len(input_controls)}) → RainForest → Output ({len(output_sensors)})")
    
    # Excel Filter Instructions
    print("\n" + "=" * 80)
    print("EXCEL FILTER INSTRUCTIONS")
    print("=" * 80)
    
    print("\n[FOR INPUT SHEET - To see control parameters:]")
    print("  1. Click filter dropdown on Column A (Physical quantity)")
    print("  2. SELECT ONLY these types:")
    print("     ✓ Temperature setpoint")
    print("     ✓ Supply fan command")
    print("     ✓ Cooling valve command")
    print("     ✓ Heating valve command")
    print("     ✓ Economizer / damper control")
    print("     ✓ Valve command / position")
    print("     ✓ Occupancy / schedule command")
    print("  3. Click filter dropdown on Column U (Biome)")
    print("     ✓ Inventory BIOM=1, BIOMNAME=Rainforest.")
    print("  4. Click OK")
    print(f"\n  Result: {len(input_controls)} climate control parameters for Rainforest")
    
    print("\n[FOR OUTPUT SHEET - To see monitoring sensors:]")
    print("  1. Click filter dropdown on Column U (Biome)")
    print("     ✓ Inventory BIOM=1, BIOMNAME=Rainforest.")
    print("  2. Column A (Physical quantity) should already show only:")
    print("     ✓ AirTempC (Temperature)")
    print("     ✓ RH (Relative Humidity)")
    print(f"\n  Result: {len(output_sensors)} climate monitoring sensors for Rainforest")
    
    print("\n" + "=" * 80)
    print("SUMMARY TABLE")
    print("=" * 80)
    print(f"\n{'Component':<30} {'Count':<15} {'Types'}")
    print("-" * 80)
    print(f"{'INPUT Parameters':<30} {len(input_controls):<15} Temperature, Fan, Valve")
    print(f"{'OUTPUT Sensors':<30} {len(output_sensors):<15} Temperature, Humidity")
    print("\n" + "=" * 80)

if __name__ == "__main__":
    main()

