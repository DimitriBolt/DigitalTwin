from __future__ import annotations

import copy
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import oracledb
from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
SENSORS_DIR = ROOT / "Sensors_Description"
WORKBOOK_PATH = SENSORS_DIR / "variables_schema.xlsx"
CONTROLS_INVENTORY_PATH = SENSORS_DIR / "Bio2-Controls-Inventory-29Jan2026.xlsx"
ENV_PATH = Path.home() / "Documents" / ".env"
ORACLE_CLIENT_LIB_DIR = Path("/opt/oracle/instantclient_19_26")
TODAY = date.today().isoformat()

CONTROL_OWNER = "BIO2CONTROLS"
CONTROL_MIRROR_OWNER = "BIO2CONTROLSALL"

HEADERS = [
    "Physical quantity / setting",
    "Physical symbol",
    "Control family",
    "Manipulated signal type",
    "Confidence",
    "Affected biome",
    "Biome/equipment note",
    "Equipment tag",
    "Equipment description / lookup",
    "BMS path (ID_NAME)",
    "Device measurement",
    "Expanded measurement / meaning",
    "Physical climate effect",
    "PDE / inverse-problem role",
    "How it enters PDE or inverse problem",
    "Source system",
    "Inventory file",
    "Inventory sheet",
    "Inventory ID",
    "Oracle owner",
    "Oracle table",
    "Oracle selector / WHERE key",
    "Time column",
    "Value column",
    "Raw unit",
    "SI / normalized unit",
    "Formula for conversion",
    "Inventory category",
    "Value facets / BACnet range",
    "Inventory variable description",
    "Availability start",
    "Availability end",
    "Value count",
    "Live-tested in Oracle?",
    "Concrete SQL query for calendar year 2025",
    "Mirror fallback",
    "Engineering verification notes",
]

BIOME_SORT = {
    "Rainforest": 0,
    "Savanna": 1,
    "Desert": 2,
    "LEO": 3,
    "Ocean": 4,
    "South lung": 5,
    "West lung": 6,
    "Wave wall": 7,
    "Orchard": 8,
    "EC": 9,
}

HARD_EXCLUDE_MEASUREMENT_TOKENS = (
    "alm",
    "fault",
    "reset",
    "runhrs",
    "errorcode",
    "battery",
    "percentload",
    "ngmeter",
    "kwh",
    "pfc",
)

CONTROL_PATTERNS = (
    r"cmd",
    r"enable$",
    r"sysenable$",
    r"blrdmd$",
    r"blrsp",
    r"effsp$",
    r"sp\d*$",
    r"spt$",
    r"vlv(cmd|out|pos)$",
    r"minvlvpos$",
    r"ed(pos|possp|cmd)$",
    r"dpr(open|close)sp$",
    r"dmpcmd$",
    r"(fan|vfd).*spd$",
    r"spdcmd$",
    r"vfd(out|cmd|dircmd)$",
    r"hertz$",
    r"rpm$",
    r"pmp(cmd|enable|psrsp|psrspt)",
    r"rainflwrat$",
    r"^sv\d+$",
    r"flamecmd$",
    r"firingrat$",
    r"oceantmpsp$",
    r"activetmpsp$",
    r"deltatmpsp$",
    r"hxsuphipsrsp$",
    r"hipsrspt$",
    r"slung(height|wtrlvl)sp$",
    r"(min|max)raintnklvlsp$",
    r"hx(sup|ret)vlvcmd$",
    r"(hw|chw)vlv(out|pos)$",
    r"rowtrvlvcmd$",
    r"muvlvcmd$",
    r"brominatorvlvcmd$",
    r"blowdnvlv",
    r"isovlvcmd$",
    r"rainvlv\d+cmd$",
    r"wtp\d+pmpcmd$",
)


@dataclass(frozen=True)
class OracleBounds:
    first_dt: datetime | None
    last_dt: datetime | None
    value_count: int | None
    error: str | None = None

    @property
    def has_data(self) -> bool:
        return self.first_dt is not None and self.last_dt is not None


def connect() -> oracledb.Connection:
    cfg = dotenv_values(ENV_PATH)
    oracledb.init_oracle_client(lib_dir=str(ORACLE_CLIENT_LIB_DIR))
    dsn = oracledb.makedsn(cfg["ORACLE_HOST"], int(cfg["ORACLE_PORT"]), sid=cfg["ORACLE_SID"])
    conn = oracledb.connect(
        user=cfg["ORACLE_USER"],
        password=cfg["ORACLE_PASSWORD"],
        dsn=dsn,
    )
    conn.call_timeout = 60_000
    return conn


def date_label(value: datetime | date | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def end_label(value: datetime | date | str | None) -> str:
    label = date_label(value)
    return "present" if label == TODAY else label


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def compact_symbol_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")


def read_controls_inventory() -> tuple[list[dict[str, Any]], dict[str, str]]:
    wb = load_workbook(CONTROLS_INVENTORY_PATH, read_only=True, data_only=True)
    summary = wb["Table Summary"]
    headers = [cell.value for cell in summary[1]]
    rows = [dict(zip(headers, values)) for values in summary.iter_rows(min_row=2, values_only=True)]

    abbreviations: dict[str, str] = {}
    for abbreviation, full_name in wb["Device Measurement Abr "].iter_rows(min_row=2, values_only=True):
        if abbreviation and full_name:
            abbreviations[str(abbreviation).strip()] = str(full_name).strip()
    return rows, abbreviations


def is_electrical_only_command(measurement: str, device_item: str, source: str) -> bool:
    measurement_l = measurement.lower()
    device_l = device_item.lower()
    source_l = source.lower()
    if any(token in measurement_l for token in ("brkclosecmd", "brkopencmd", "brkaclosecmd", "brkaopencmd")):
        return True
    electrical_devices = ("lc", "sg", "gen")
    if measurement in {"CloseCmd", "OpenCmd", "SelectorCmd", "EStopCmd"}:
        return device_l.startswith(electrical_devices) or source_l.startswith(electrical_devices)
    return False


def include_control_candidate(row: dict[str, Any]) -> bool:
    measurement = normalize_text(row.get("Device_Measurement"))
    measurement_l = measurement.lower()
    if not measurement:
        return False
    if any(token in measurement_l for token in HARD_EXCLUDE_MEASUREMENT_TOKENS):
        return False
    if "spcond" in measurement_l:
        return False
    if measurement_l.endswith("sts") and not any(
        token in measurement_l for token in ("cmd", "sp", "spt", "pos", "out", "enable", "dmd")
    ):
        return False
    if measurement_l.endswith("sw") or "doorsw" in measurement_l or "stopsw" in measurement_l:
        return False
    if is_electrical_only_command(
        measurement,
        normalize_text(row.get("Device_Item")),
        normalize_text(row.get("SOURCE")),
    ):
        return False
    return any(re.search(pattern, measurement_l) for pattern in CONTROL_PATTERNS)


def control_family(measurement: str, device_item: str) -> str:
    name = measurement.lower()
    device = device_item.upper()
    if re.match(r"^sv\d+$", measurement, re.IGNORECASE):
        return "Solenoid valve / unknown valve"
    if "rain" in name or "tesco" in name or "irrig" in name:
        return "Rain / irrigation control"
    if "ccvlvcmd" in name:
        return "Cooling valve command"
    if "hcvlvcmd" in name:
        return "Heating valve command"
    if "twccvlvcmd" in name:
        return "Tower-water control valve command"
    if "isovlvcmd" in name:
        return "Isolation valve command"
    if "vlvcmd" in name or "vlvout" in name or "vlvpos" in name or "minvlvpos" in name:
        return "Valve command / position"
    if "edcmd" in name or "edpos" in name:
        return "Economizer / damper control"
    if "dmp" in name or "dpr" in name:
        return "Damper control"
    if "sfcmd" in name:
        return "Supply fan command"
    if "fan" in name or "vfd" in name or "spd" in name or "hertz" in name or "rpm" in name:
        return "Fan / VFD speed control"
    if "satmpsp" in name or "tmpsp" in name:
        return "Temperature setpoint"
    if "psrsp" in name or "psrspt" in name or "hipsrspt" in name:
        return "Pressure setpoint"
    if "lvlsp" in name or "heightsp" in name:
        return "Level / height setpoint"
    if "pmp" in name or device.startswith(("CHWP", "HWP", "TWP", "FWP")):
        return "Pump command"
    if "chlr" in name:
        return "Chiller command / setpoint"
    if "effsp" in name:
        return "Boiler efficiency setpoint"
    if "blr" in name:
        return "Boiler command / setpoint"
    if "fwdcmd" in name or "revcmd" in name:
        return "Fan / VFD direction command"
    if "exh" in name and "cmd" in name:
        return "Exhaust fan command"
    if "runcmd" in name and device.startswith("COMPC"):
        return "Compressor run command"
    if name == "cmd" and device.startswith("CV"):
        return "Valve command / position"
    if "occ" in name:
        return "Occupancy / schedule command"
    if "enable" in name:
        return "System enable"
    return "Control parameter / needs classification"


def signal_type(measurement: str) -> str:
    name = measurement.lower()
    if re.match(r"^sv\d+$", measurement, re.IGNORECASE):
        return "binary valve command or status/command point"
    if "cmd" in name or name.endswith("enable"):
        return "command"
    if re.search(r"(sp|spt)\d*$", name):
        return "setpoint"
    if "pos" in name or "out" in name or "spd" in name or "hertz" in name or "rpm" in name:
        return "actuator output / position feedback"
    if "dmd" in name:
        return "controller demand"
    return "needs engineering verification"


def confidence(row: dict[str, Any], family: str, sig_type: str) -> str:
    measurement = normalize_text(row.get("Device_Measurement"))
    measurement_l = measurement.lower()
    category = normalize_text(row.get("Category")).lower()
    biome = normalize_text(row.get("BIOMNAME"))
    if biome == "EC":
        return "Medium - Energy Center downstream biome unresolved"
    if "status" in category and "command" not in category and "cmd" not in measurement_l:
        return "Low - inventory category is status; verify writeability"
    if "actuator output" in sig_type or "feedback" in sig_type:
        return "Medium - actuator output/feedback, not necessarily operator-entered"
    if "unknown" in family.lower() or "needs" in sig_type:
        return "Low - equipment meaning requires engineering verification"
    if measurement_l.startswith("sv"):
        return "Medium - likely solenoid valve, purpose not described in inventory"
    return "High - name is command/setpoint for named climate equipment"


def equipment_description(device_item: str, biome: str) -> str:
    device = device_item.upper()
    if device.startswith("AHUR"):
        return f"{biome} air-handling unit {device_item}; use equipment tag and BMS path to find the physical AHU."
    if device.startswith("AHUS"):
        return f"{biome} air-handling unit {device_item}; use equipment tag and BMS path to find the physical AHU."
    if device.startswith("AHUD"):
        return f"{biome} air-handling unit {device_item}; use equipment tag and BMS path to find the physical AHU."
    if device.startswith("AHUL"):
        return f"{biome} air-handling unit {device_item}; use equipment tag and BMS path to find the physical AHU."
    if device.startswith("MISCRF"):
        return "Rainforest miscellaneous climate controls; verify exact field device from BMS path."
    if device.startswith("MISCR1"):
        return "Rainforest damper/control-sensor setpoint group; verify exact field device from BMS path."
    if device.startswith("MISCSAV") and biome == "Rainforest":
        return "Rainforest rain/Tesco subsystem recorded under MiscSAV tag; verify exact equipment in BMS."
    if device.startswith("MISCSAV"):
        return f"{biome} miscellaneous climate controls; verify exact field device from BMS path."
    if device.startswith("MISCDES"):
        return f"{biome} miscellaneous climate controls; verify exact field device from BMS path."
    if device == "HEX01":
        return "Ocean heat exchanger HEX01."
    if device.startswith("BLR") or re.match(r"F1A\d+", device):
        return "Energy Center boiler/hot-water plant equipment; downstream biome not resolved in inventory."
    if device.startswith("CHLR"):
        return "Energy Center chiller/cold-water plant equipment; downstream biome not resolved in inventory."
    if device.startswith(("CHWP", "HWP", "TWP", "FWP")):
        return "Energy Center water-loop pump; downstream loop/biome should be verified by engineers."
    if device.startswith(("CHW", "CV", "SVT", "TW")):
        return "Energy Center hydronic/tower-water valve or loop equipment; downstream loop should be verified."
    if device.startswith(("VFD", "CTH")):
        return "Energy Center VFD or cooling-tower fan equipment; downstream loop should be verified."
    if device.startswith("COMPC"):
        return "Energy Center compressor equipment; downstream function should be verified."
    if device == "WAVEWALL":
        return "Wave wall controls."
    return f"{biome} equipment tag {device_item}; use BMS path for physical lookup."


def biome_note(row: dict[str, Any]) -> str:
    biome = normalize_text(row.get("BIOMNAME"))
    biom_id = normalize_text(row.get("BIOM"))
    if biome == "EC":
        return "Inventory BIOMNAME=EC. This is central Energy Center equipment; affected biome is plant-wide or unresolved."
    return f"Inventory BIOM={biom_id}, BIOMNAME={biome}."


def physical_effect(family: str, biome: str) -> str:
    family_l = family.lower()
    if "cooling" in family_l or "chiller" in family_l:
        return f"Changes cooling capacity available to {biome} or to the shared chilled-water loop."
    if "heating" in family_l or "boiler" in family_l:
        return f"Changes heating capacity available to {biome} or to the shared hot-water loop."
    if "fan" in family_l or "vfd" in family_l:
        return f"Changes airflow, exchange, pressure balance, or mixing for {biome}."
    if "damper" in family_l or "economizer" in family_l:
        return f"Changes ventilation path, outside-air fraction, or air exchange for {biome}."
    if "rain" in family_l or "irrigation" in family_l:
        return f"Changes water input or irrigation forcing for {biome}."
    if "pump" in family_l:
        return f"Changes water-loop circulation or pressure serving {biome} or a shared plant loop."
    if "valve" in family_l:
        return f"Changes hydronic/air/water flow through equipment serving {biome}."
    if "temperature setpoint" in family_l:
        return f"Sets target temperature used by equipment serving {biome}."
    if "pressure setpoint" in family_l:
        return f"Sets target pressure used by equipment serving {biome}."
    if "level" in family_l:
        return f"Sets water level or height target for {biome} support equipment."
    return f"Potential control input affecting {biome}; physical pathway needs verification."


def pde_role(family: str, sig_type: str) -> tuple[str, str]:
    family_l = family.lower()
    if "rain" in family_l or "irrigation" in family_l:
        return (
            "Source-term control parameter",
            "Use as prescribed water-input/source forcing or as a candidate manipulated variable in control/inversion.",
        )
    if "temperature setpoint" in family_l or "cooling" in family_l or "heating" in family_l:
        return (
            "Thermal boundary-control parameter",
            "Use as known actuator trajectory or estimate an actuator-to-boundary map for heat flux/air temperature.",
        )
    if "fan" in family_l or "damper" in family_l or "economizer" in family_l:
        return (
            "Air-exchange boundary-control parameter",
            "Use as known actuator trajectory for ventilation, pressure exchange, or mixing boundary conditions.",
        )
    if "pump" in family_l or "valve" in family_l or "pressure" in family_l or "level" in family_l:
        return (
            "Hydraulic / plant-support control parameter",
            "Use as plant forcing or candidate explanatory variable for water, heat, and air-handling boundary behavior.",
        )
    if sig_type == "actuator output / position feedback":
        return (
            "Actuator-response observable",
            "Use to identify the mapping between operator command, equipment response, and resulting climate boundary.",
        )
    return (
        "Control parameter requiring verification",
        "Retain for engineering review; include in models only after confirming physical actuator pathway.",
    )


def expanded_measurement(measurement: str, abbreviations: dict[str, str]) -> str:
    if measurement in abbreviations:
        return abbreviations[measurement]
    pieces = re.sub(r"([a-z])([A-Z])", r"\1 \2", measurement).replace("_", " ")
    exact_terms = []
    for key, value in sorted(abbreviations.items(), key=lambda item: len(item[0]), reverse=True):
        if key and key in measurement and len(key) > 2:
            exact_terms.append(f"{key}={value}")
    if exact_terms:
        return f"{pieces}; " + "; ".join(exact_terms[:4])
    return pieces


def conversion(unit: Any, variable: Any) -> tuple[str, str, str, str]:
    raw_unit = normalize_text(unit)
    variable_text = normalize_text(variable)
    if not raw_unit:
        return "", "raw state", "value_si = raw VALUE", "t.VALUE"
    if raw_unit == "%":
        return raw_unit, "fraction", "value_si = VALUE / 100", "t.VALUE / 100"
    if raw_unit in {"degF", "F"} or "F" in raw_unit:
        return raw_unit, "K", "T_K = (VALUE - 32) * 5 / 9 + 273.15", "(t.VALUE - 32) * 5 / 9 + 273.15"
    if raw_unit == "L/min":
        return raw_unit, "m^3/s", "Q = VALUE * 1e-3 / 60", "t.VALUE * 1e-3 / 60"
    if raw_unit == "psi":
        return raw_unit, "Pa", "p_Pa = VALUE * 6894.757", "t.VALUE * 6894.757"
    if raw_unit == "in":
        return raw_unit, "m", "h_m = VALUE * 0.0254", "t.VALUE * 0.0254"
    if raw_unit == "inH2O" and "level" in variable_text.lower():
        return raw_unit, "m water", "h_m = VALUE * 0.0254", "t.VALUE * 0.0254"
    if raw_unit == "inH2O":
        return raw_unit, "Pa", "p_Pa = VALUE * 249.0889", "t.VALUE * 249.0889"
    return raw_unit, raw_unit, "value_si = raw VALUE", "t.VALUE"


def make_query(owner: str, table_name: str, value_expr: str) -> str:
    return (
        "SELECT\n"
        "    t.TIMESTAMP,\n"
        "    t.VALUE AS raw_value,\n"
        f"    {value_expr} AS value_si\n"
        f"FROM {owner}.{table_name} t\n"
        "WHERE t.TIMESTAMP >= DATE '2025-01-01'\n"
        "  AND t.TIMESTAMP <  DATE '2026-01-01'\n"
        "ORDER BY t.TIMESTAMP;"
    )


def fetch_existing_tables(cur: oracledb.Cursor, owners: tuple[str, ...]) -> dict[str, set[str]]:
    placeholders = ", ".join(f":owner{i}" for i, _ in enumerate(owners))
    params = {f"owner{i}": owner for i, owner in enumerate(owners)}
    cur.execute(
        f"""
        SELECT owner, table_name
        FROM all_tables
        WHERE owner IN ({placeholders})
        """,
        params,
    )
    result = {owner: set() for owner in owners}
    for owner, table_name in cur.fetchall():
        result.setdefault(str(owner), set()).add(str(table_name))
    return result


def fetch_bounds(cur: oracledb.Cursor, owner: str, table_name: str) -> OracleBounds:
    try:
        cur.execute(
            f"""
            SELECT MIN(t.TIMESTAMP), MAX(t.TIMESTAMP), COUNT(*)
            FROM {owner}.{table_name} t
            """
        )
        first_dt, last_dt, value_count = cur.fetchone()
        return OracleBounds(first_dt, last_dt, int(value_count or 0))
    except Exception as exc:  # noqa: BLE001 - record validation failure in workbook.
        return OracleBounds(None, None, None, str(exc).splitlines()[0])


def choose_owner(table_name: str, existing_tables: dict[str, set[str]]) -> str:
    if table_name in existing_tables.get(CONTROL_OWNER, set()):
        return CONTROL_OWNER
    if table_name in existing_tables.get(CONTROL_MIRROR_OWNER, set()):
        return CONTROL_MIRROR_OWNER
    return CONTROL_OWNER


def validation_text(owner: str, table_name: str, bounds: OracleBounds, existing_tables: dict[str, set[str]]) -> str:
    if bounds.has_data:
        return f"Yes - {owner}.{table_name} live-tested in Oracle on {TODAY}."
    if table_name not in existing_tables.get(owner, set()):
        return f"No - {owner}.{table_name} not found in ALL_TABLES on {TODAY}."
    return f"No - table found, but bounds query failed or returned no rows on {TODAY}: {bounds.error or 'no rows'}"


def verification_notes(row: dict[str, Any], family: str, confidence_value: str, owner: str, bounds: OracleBounds) -> str:
    notes = [
        "Candidate selected from Bio2 Controls inventory because the point name looks like a command, setpoint, actuator output, valve/damper/fan/pump speed, or irrigation handle.",
        f"Selection family: {family}.",
    ]
    if "Medium" in confidence_value or "Low" in confidence_value:
        notes.append(confidence_value)
    if normalize_text(row.get("BIOMNAME")) == "EC":
        notes.append("Engineering review should map this central plant point to downstream loops/biomes.")
    category = normalize_text(row.get("Category"))
    if category and "SENSOR" in category and any(token in normalize_text(row.get("Device_Measurement")).lower() for token in ("cmd", "sp", "pos", "out", "spd")):
        notes.append("Inventory category says SENSOR, but point name is control-like; verify whether it is command, setpoint, or actuator feedback.")
    if bounds.error:
        notes.append(f"Oracle validation error: {bounds.error}")
    return " ".join(notes)


def build_records(cur: oracledb.Cursor | None = None) -> list[dict[str, Any]]:
    rows, abbreviations = read_controls_inventory()
    selected = [row for row in rows if include_control_candidate(row)]
    existing_tables: dict[str, set[str]] = {}
    bounds_by_table: dict[tuple[str, str], OracleBounds] = {}
    if cur is not None:
        existing_tables = fetch_existing_tables(cur, (CONTROL_OWNER, CONTROL_MIRROR_OWNER))

    records: list[dict[str, Any]] = []
    for index, row in enumerate(selected, start=1):
        table_name = normalize_text(row.get("TABLE_NAME"))
        owner = choose_owner(table_name, existing_tables) if existing_tables else CONTROL_OWNER
        key = (owner, table_name)
        if cur is not None and key not in bounds_by_table:
            bounds_by_table[key] = fetch_bounds(cur, owner, table_name)
            if index % 50 == 0:
                print(f"validated {index}/{len(selected)} control tables")
        bounds = bounds_by_table.get(key, OracleBounds(row.get("STARTDATE"), row.get("ENDDATE"), row.get("VALUECNT")))

        measurement = normalize_text(row.get("Device_Measurement"))
        device_item = normalize_text(row.get("Device_Item"))
        biome = normalize_text(row.get("BIOMNAME"))
        family = control_family(measurement, device_item)
        sig_type = signal_type(measurement)
        conf = confidence(row, family, sig_type)
        raw_unit, si_unit, formula, value_expr = conversion(row.get("UNITS"), row.get("VARIABLES"))
        role, enters = pde_role(family, sig_type)

        first = date_label(bounds.first_dt) or date_label(row.get("STARTDATE"))
        last = end_label(bounds.last_dt) or end_label(row.get("ENDDATE"))
        value_count = bounds.value_count if bounds.value_count is not None else row.get("VALUECNT")
        source_system = "Bio2 Controls"
        record = {
            "Physical quantity / setting": family,
            "Physical symbol": "u_" + compact_symbol_part(f"{device_item}_{measurement}"),
            "Control family": family,
            "Manipulated signal type": sig_type,
            "Confidence": conf,
            "Affected biome": biome or "unknown",
            "Biome/equipment note": biome_note(row),
            "Equipment tag": device_item,
            "Equipment description / lookup": equipment_description(device_item, biome),
            "BMS path (ID_NAME)": normalize_text(row.get("ID_NAME")),
            "Device measurement": measurement,
            "Expanded measurement / meaning": expanded_measurement(measurement, abbreviations),
            "Physical climate effect": physical_effect(family, biome or "unknown biome"),
            "PDE / inverse-problem role": role,
            "How it enters PDE or inverse problem": enters,
            "Source system": source_system,
            "Inventory file": CONTROLS_INVENTORY_PATH.name,
            "Inventory sheet": "Table Summary",
            "Inventory ID": row.get("ID"),
            "Oracle owner": owner,
            "Oracle table": table_name,
            "Oracle selector / WHERE key": "WHERE TIMESTAMP BETWEEN :t0 AND :t1 ORDER BY TIMESTAMP",
            "Time column": "TIMESTAMP",
            "Value column": "VALUE",
            "Raw unit": raw_unit,
            "SI / normalized unit": si_unit,
            "Formula for conversion": formula,
            "Inventory category": normalize_text(row.get("Category")),
            "Value facets / BACnet range": normalize_text(row.get("VALUEFACETS")),
            "Inventory variable description": normalize_text(row.get("VARIABLES")),
            "Availability start": first,
            "Availability end": last,
            "Value count": value_count,
            "Live-tested in Oracle?": validation_text(owner, table_name, bounds, existing_tables) if existing_tables else "",
            "Concrete SQL query for calendar year 2025": make_query(owner, table_name, value_expr),
            "Mirror fallback": f"{CONTROL_MIRROR_OWNER}.{table_name}",
            "Engineering verification notes": verification_notes(row, family, conf, owner, bounds),
        }
        records.append(record)

    records.sort(
        key=lambda rec: (
            BIOME_SORT.get(str(rec["Affected biome"]), 99),
            str(rec["Equipment tag"]),
            str(rec["Control family"]),
            str(rec["Device measurement"]),
            int(rec["Inventory ID"] or 0),
        )
    )
    return records


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        src = ws.cell(source_row, col_idx)
        dst = ws.cell(target_row, col_idx)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)


def write_input_sheet(records: list[dict[str, Any]]) -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Input"]
    max_col = len(HEADERS)
    max_row = max(ws.max_row, len(records) + 3)

    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max(ws.max_column, max_col)):
        for cell in row:
            cell.value = None

    ws["A1"] = "Bio2 climate-equipment control inputs"
    ws["A2"] = (
        "Generated from Bio2-Controls-Inventory-29Jan2026.xlsx. "
        "Rows are candidate manipulated controls, setpoints, and actuator outputs; uncertain equipment mappings are flagged for engineering review."
    )
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(3, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, record in enumerate(records, start=4):
        copy_row_style(ws, 4, row_idx, max_col)
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row_idx, col_idx, record.get(header))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx in range(len(records) + 4, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None

    widths = {
        "A": 28,
        "B": 28,
        "C": 24,
        "D": 24,
        "E": 28,
        "F": 14,
        "G": 34,
        "H": 16,
        "I": 44,
        "J": 40,
        "K": 22,
        "L": 42,
        "M": 46,
        "N": 28,
        "O": 46,
        "P": 18,
        "Q": 32,
        "R": 18,
        "S": 12,
        "T": 18,
        "U": 40,
        "V": 36,
        "W": 14,
        "X": 14,
        "Y": 14,
        "Z": 18,
        "AA": 28,
        "AB": 24,
        "AC": 40,
        "AD": 26,
        "AE": 16,
        "AF": 16,
        "AG": 14,
        "AH": 42,
        "AI": 70,
        "AJ": 42,
        "AK": 70,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{len(records) + 3}"
    ws.sheet_view.showGridLines = True

    wb.save(WORKBOOK_PATH)


def main() -> None:
    conn = connect()
    cur = conn.cursor()
    try:
        records = build_records(cur)
        write_input_sheet(records)
        print(f"Wrote {len(records)} Input control rows to {WORKBOOK_PATH}")
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
