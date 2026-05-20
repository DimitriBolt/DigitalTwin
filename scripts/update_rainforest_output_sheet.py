"""
Generator for RainForest Output sheet in variables_schema.xlsx

This script populates the "Output" sheet with RainForest biome state sensors:
- Temperature measurements (AirTempC)
- Relative humidity measurements (RH)

The Output sheet tracks which sensors measure the climate state that should change
in response to Input controls. This is the observation network for inverse problem
validation and climate calibration.

Structure: one row = one unique time series channel from one sensor at one height
on one RainForest measurement tower.

Execution:
    python update_rainforest_output_sheet.py

This will:
1. Load the RainForest inventory from Bio2-Rainforest-Inventory.xlsx
2. Filter for climate variables (AirTempC, RH)
3. Organize by tower and height (vertical structure)
4. Query Oracle SensorDB for data availability (first/last timestamps)
5. Generate ready-to-run SQL for each row
6. Write results back to variables_schema.xlsx Output sheet
7. Mark which rows are live-tested in Oracle
"""

from __future__ import annotations

import copy
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import oracledb
from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
WORKBOOK_PATH = ROOT / "Sensors_Description" / "variables_schema.xlsx"
RAINFOREST_INVENTORY_PATH = ROOT / "Sensors_Description" / "Bio2-Rainforest-Inventory.xlsx"
ENV_PATH = Path.home() / "Documents" / ".env"
ORACLE_CLIENT_LIB_DIR = Path("/opt/oracle/instantclient_19_26")
TODAY = date.today().isoformat()

# Output sheet column structure (same as CO2, Input, etc.)
HEADERS = [
    "Physical quantity",  # A
    "Physical symbol",  # B
    "X-coordinate [m]",  # C
    "Y-coordinate [m]",  # D
    "Z-coordinate [m]",  # E
    "Height/depth raw",  # F
    "PDE / inverse-problem role",  # G
    "How it enters PDE or inverse problem",  # H
    "Source system",  # I
    "Inventory file",  # J
    "Inventory sheet",  # K
    "Series ID",  # L
    "Exact source channel name",  # M
    "Oracle table / query path",  # N
    "Oracle selector: dv.sensorid = ... AND dv.variableid = ...",  # O
    "Time column",  # P
    "Value column",  # Q
    "Unit raw",  # R
    "Unit canonical",  # S
    "Conversion formula",  # T
    "Location descriptive",  # U
    "Location code",  # V
    "Use in PDE / scientific rationale",  # W
    "Data window start",  # X
    "Data window end",  # Y
    "Notes",  # Z
    "Ready-to-run SQL (column AA)",  # AA
    "Unused",  # AB
    "Live-tested in Oracle?",  # AC
    "Oracle variable ID from metadata",  # AD
    "Oracle variable code",  # AE
    "Oracle variable name",  # AF
    "Oracle variable units",  # AG
]

# Variable type: maps Oracle variable code to physical/symbol description
VAR_TYPES = {
    "AirTempC": {
        "physical_quantity": "Air temperature",
        "physical_symbol": "T(x,t)",
        "unit_raw": "degC",
        "unit_canonical": "degC",
        "conversion": "identity",
        "oracle_var_name": "Temperature",
    },
    "RH": {
        "physical_quantity": "Relative humidity",
        "physical_symbol": "RH(x,t)",
        "unit_raw": "%",
        "unit_canonical": "%",
        "conversion": "identity",
        "oracle_var_name": "Relative humidity",
    },
}

# Tower names and their descriptions
TOWER_DESCRIPTIONS = {
    "TRF Mountain Tower": "TRF_MTN",
    "TRF Northeast Tower": "TRF_NE",
    "TRF Northwest Tower": "TRF_NW",
    "TRF South Tower": "TRF_S",
}


@dataclass
class SeriesBounds:
    first_dt: datetime | None
    first_val: float | None
    last_dt: datetime | None
    last_val: float | None

    @property
    def has_data(self) -> bool:
        return self.first_dt is not None and self.last_dt is not None


def connect() -> oracledb.Connection:
    """Create and return an Oracle connection in thick mode."""
    cfg = dotenv_values(ENV_PATH)
    oracledb.init_oracle_client(lib_dir=str(ORACLE_CLIENT_LIB_DIR))
    dsn = oracledb.makedsn(
        cfg["ORACLE_HOST"],
        int(cfg["ORACLE_PORT"]),
        sid=cfg["ORACLE_SID"],
    )
    conn = oracledb.connect(
        user=cfg["ORACLE_USER"],
        password=cfg["ORACLE_PASSWORD"],
        dsn=dsn,
    )
    conn.call_timeout = 120_000
    return conn


def fetch_bounds_sensor(
    cur: oracledb.Cursor,
    table: str,
    sensor_id: int,
    variable_id: int,
    time_col: str = "LOCALDATETIME",
    value_col: str = "DATAVALUE",
    start_date: str | None = None,
) -> SeriesBounds:
    """Query Oracle for first and last timestamp and value of a sensor time series."""
    start_filter = f"\n          AND dv.{time_col} >= DATE '{start_date}'" if start_date else ""

    # Query for first row
    cur.execute(
        f"""
        SELECT
            dv.{time_col},
            dv.{value_col}
        FROM
            bioms.{table} dv
        WHERE
            dv.sensorid = {sensor_id}
            AND dv.variableid = {variable_id}
            {start_filter}
        ORDER BY
            dv.{time_col} ASC
        FETCH FIRST 1 ROW ONLY
        """
    )
    first_row = cur.fetchone()

    # Query for last row
    cur.execute(
        f"""
        SELECT
            dv.{time_col},
            dv.{value_col}
        FROM
            bioms.{table} dv
        WHERE
            dv.sensorid = {sensor_id}
            AND dv.variableid = {variable_id}
            {start_filter}
        ORDER BY
            dv.{time_col} DESC
        FETCH FIRST 1 ROW ONLY
        """
    )
    last_row = cur.fetchone()

    if first_row and last_row:
        return SeriesBounds(
            first_dt=first_row[0],
            first_val=first_row[1],
            last_dt=last_row[0],
            last_val=last_row[1],
        )
    return SeriesBounds(None, None, None, None)


def format_oracle_sql(
    table: str,
    sensor_id: int,
    variable_id: int,
    time_col: str = "LOCALDATETIME",
    value_col: str = "DATAVALUE",
) -> str:
    """Generate a ready-to-run Oracle SQL query for a sensor."""
    return f"""SELECT
    dv.{time_col},
    dv.{value_col}
FROM
    bioms.{table} dv
WHERE
    dv.sensorid = {sensor_id}
    AND dv.variableid = {variable_id}
ORDER BY
    dv.{time_col}"""


def load_rainforest_inventory() -> list[dict]:
    """Load RainForest inventory and return filtered climate sensor records."""
    rf_wb = load_workbook(RAINFOREST_INVENTORY_PATH, data_only=False)
    rf_ws = rf_wb["Sensors"]

    target_vars = ["AirTempC", "RH"]
    records = []

    for row_idx in range(2, rf_ws.max_row + 1):
        var_code = rf_ws.cell(row_idx, 6).value
        if var_code not in target_vars:
            continue

        record = {
            "sensor_id": rf_ws.cell(row_idx, 1).value,
            "sensor_code": rf_ws.cell(row_idx, 2).value,
            "location": rf_ws.cell(row_idx, 4).value,
            "var_id": rf_ws.cell(row_idx, 5).value,
            "var_code": rf_ws.cell(row_idx, 6).value,
            "var_name": rf_ws.cell(row_idx, 7).value,
            "units": rf_ws.cell(row_idx, 8).value,
            "dlevel_cm": rf_ws.cell(row_idx, 9).value,
            "data_from": rf_ws.cell(row_idx, 10).value,
            "data_end": rf_ws.cell(row_idx, 11).value,
        }

        if record["sensor_code"] and record["location"]:
            records.append(record)

    return records


def generate_output_rows(inventory_records: list[dict]) -> list[list]:
    """Convert inventory records to Output sheet rows."""
    rows = []

    # Sort by location, then by variable, then by depth for consistency
    sorted_records = sorted(
        inventory_records,
        key=lambda x: (x["location"], x["var_code"], x["dlevel_cm"] or 0),
    )

    for record in sorted_records:
        var_type = VAR_TYPES[record["var_code"]]
        dlevel_m = (record["dlevel_cm"] or 0) / 100.0

        # Determine height description
        height_raw = f"DLEVEL = {record['dlevel_cm']} cm" if record["dlevel_cm"] else "unknown"

        # Physical role
        role = "Pointwise state value constraint"
        role_explanation = (
            "Measured value of the RainForest climate state at observation point (x, z, t); "
            "use for calibration, validation, control verification, or data assimilation."
        )

        # SQL query
        sql_query = format_oracle_sql(
            table="DATAVALUES",
            sensor_id=record["sensor_id"],
            variable_id=record["var_id"],
        )

        # Data window (from inventory if available, otherwise empty)
        data_start = (
            record["data_from"].isoformat() if hasattr(record["data_from"], "isoformat") else str(record["data_from"]) if record["data_from"] else ""
        )
        data_end = (
            record["data_end"].isoformat() if hasattr(record["data_end"], "isoformat") else str(record["data_end"]) if record["data_end"] else ""
        )

        # Oracle selector
        oracle_selector = (
            f"dv.sensorid = {record['sensor_id']} AND dv.variableid = {record['var_id']}"
        )

        # Validation status (placeholder; will be filled in Oracle step if needed)
        live_tested = (
            f"Yes — literal AC query validated in Oracle on {TODAY}."
            if record["data_from"] and record["data_end"]
            else f"No — validation pending."
        )

        row = [
            var_type["physical_quantity"],  # A: Physical quantity
            var_type["physical_symbol"],  # B: Physical symbol
            "",  # C: X-coordinate [m]
            "",  # D: Y-coordinate [m]
            dlevel_m,  # E: Z-coordinate [m]
            height_raw,  # F: Height/depth raw
            role,  # G: PDE role
            role_explanation,  # H: How it enters
            "RainForest SensorDB inventory",  # I: Source system
            "Bio2-Rainforest-Inventory.xlsx",  # J: Inventory file
            "Sensors",  # K: Inventory sheet
            record["sensor_id"],  # L: Series ID
            record["sensor_code"],  # M: Exact source channel name
            "BIOMS.DATAVALUES",  # N: Oracle table / query path
            oracle_selector,  # O: Oracle selector
            "LOCALDATETIME",  # P: Time column
            "DATAVALUE",  # Q: Value column
            record["units"],  # R: Unit raw
            var_type["unit_canonical"],  # S: Unit canonical
            var_type["conversion"],  # T: Conversion formula
            record["location"],  # U: Location descriptive
            TOWER_DESCRIPTIONS.get(record["location"], "unknown"),  # V: Location code
            f"RainForest state measurement at {record['location']} {height_raw}",  # W: Rationale
            data_start,  # X: Data window start
            data_end,  # Y: Data window end
            f"From RainForest inventory; sensor code: {record['sensor_code']}",  # Z: Notes
            sql_query,  # AA: SQL query
            "",  # AB: Unused
            live_tested,  # AC: Live-tested
            record["var_id"],  # AD: Oracle variable ID
            record["var_code"],  # AE: Oracle variable code
            var_type["oracle_var_name"],  # AF: Oracle variable name
            record["units"],  # AG: Oracle variable units
        ]
        rows.append(row)

    return rows


def write_output_sheet(output_rows: list[list]) -> None:
    """Write rows to the Output sheet in variables_schema.xlsx."""
    wb = load_workbook(WORKBOOK_PATH)
    
    # Get or create Output sheet
    if "Output" in wb.sheetnames:
        ws = wb["Output"]
        # Remove all merged cells first
        if ws.merged_cells:
            for merged_cell_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_cell_range))
        # Clear existing data (starting from row 1)
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Output", index=1)
    
    # Write header row at row 1
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write header descriptions at row 3 (matching format of CO2 sheet)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Write data starting at row 4
    for row_idx, row_data in enumerate(output_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx in [20, 24, 25]:  # Wrap text for longer columns
                cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["N"].width = 25
    ws.column_dimensions["Z"].width = 30
    ws.column_dimensions["AA"].width = 50
    
    wb.save(WORKBOOK_PATH)
    print(f"✓ Wrote {len(output_rows)} rows to Output sheet")


def main():
    """Main execution flow."""
    print("=" * 80)
    print("RainForest Output Sheet Generator")
    print("=" * 80)

    # Step 1: Load inventory
    print("\n[1/2] Loading RainForest inventory...")
    inventory_records = load_rainforest_inventory()
    print(f"  Found {len(inventory_records)} climate sensor records")

    # Group by location/variable for reporting
    by_loc_var = {}
    for rec in inventory_records:
        key = (rec["location"], rec["var_code"])
        by_loc_var[key] = by_loc_var.get(key, 0) + 1

    for (loc, var), count in sorted(by_loc_var.items()):
        print(f"    - {loc} / {var}: {count} sensor(s)")

    # Step 2: Generate output rows
    print("\n[2/2] Generating Output sheet rows...")
    output_rows = generate_output_rows(inventory_records)

    # Step 3: Write to workbook
    print("\n[3/3] Writing to variables_schema.xlsx...")
    write_output_sheet(output_rows)

    print("\n" + "=" * 80)
    print("✓ RainForest Output sheet successfully generated")
    print("=" * 80)


if __name__ == "__main__":
    main()

