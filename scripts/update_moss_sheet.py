from __future__ import annotations

import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import oracledb
from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
SENSORS_DIR = ROOT / "Sensors_Description"
WORKBOOK_PATH = SENSORS_DIR / "variables_schema.xlsx"
ENV_PATH = Path.home() / "Documents" / ".env"
ORACLE_CLIENT_LIB_DIR = Path("/opt/oracle/instantclient_19_26")
TODAY = date.today().isoformat()

SLOPES = [
    {
        "label": "LEO East",
        "schema": "LEO_EAST",
        "schema_prefix": "leo_east",
        "inventory": SENSORS_DIR / "LEO-East-Inventory.xlsx",
    },
    {
        "label": "LEO Center",
        "schema": "LEO_CENTER",
        "schema_prefix": "leo_center",
        "inventory": SENSORS_DIR / "LEO-Center-Inventory.xlsx",
    },
    {
        "label": "LEO West",
        "schema": "LEO_WEST",
        "schema_prefix": "leo_west",
        "inventory": SENSORS_DIR / "LEO-West-Inventory.xlsx",
    },
]

KEEP_VARIABLE_NAMES = {
    "Temperature",
    "Relative humidity",
    "Volumetric water content",
    "Water potential",
    "Water vapor concentration",
}
HUMIDITY_VARIABLE_CODES = {
    "VWC",
    "MWP",
    "RH",
    "H2O_cellA",
    "H2O_cellB",
}

AIR_LEVEL_HEIGHT_M = {
    1: 0.25,
    2: 1.0,
    3: 3.0,
    4: 6.0,
    5: 10.0,
}

YES_VALIDATED = f"Yes - literal AA query validated in Oracle on {TODAY}."
NO_ZERO_ROWS = f"No - literal AA query returns zero rows in current Oracle export as of {TODAY}."


@dataclass(frozen=True)
class OracleSensorInfo:
    sensor_code: str | None
    data_table_name: str | None
    location_name: str | None
    local_x: float | None
    local_y: float | None
    local_z: float | None
    box_z: float | None
    box_vertical_datum: str | None
    dlevel: str | None
    elevation: float | None


@dataclass(frozen=True)
class OracleVariableInfo:
    variable_id: int
    variable_code: str | None
    variable_name: str | None
    units: str | None


@dataclass(frozen=True)
class CatalogBounds:
    first_day: str
    last_day: str
    value_count: int


@dataclass(frozen=True)
class SeriesBounds:
    first_dt: datetime | None
    first_val: float | None
    last_dt: datetime | None
    last_val: float | None
    validation_source: str = "literal AA query"
    validation_error: str | None = None

    @property
    def has_data(self) -> bool:
        return self.first_dt is not None and self.last_dt is not None


def connect() -> oracledb.Connection:
    cfg = dotenv_values(ENV_PATH)
    oracledb.init_oracle_client(lib_dir=str(ORACLE_CLIENT_LIB_DIR))
    dsn = oracledb.makedsn(cfg["ORACLE_HOST"], int(cfg["ORACLE_PORT"]), sid=cfg["ORACLE_SID"])
    conn = oracledb.connect(
        user=cfg["ORACLE_USER"],
        password=cfg["ORACLE_PASSWORD"],
        dsn=dsn,
    )
    conn.call_timeout = 30_000
    return conn


def normalize_header(value: Any) -> str:
    return str(value or "").strip().casefold()


def header_indexes(header: tuple[Any, ...]) -> dict[str, list[int]]:
    result: dict[str, list[int]] = {}
    for idx, value in enumerate(header):
        key = normalize_header(value)
        if key:
            result.setdefault(key, []).append(idx)
    return result


def get_index(indexes: dict[str, list[int]], *names: str, last: bool = False) -> int | None:
    for name in names:
        values = indexes.get(normalize_header(name))
        if values:
            return values[-1] if last else values[0]
    return None


def row_value(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def date_label(value: datetime | date | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if str(value).strip().lower() == "present":
        return "present"
    return str(value).strip()


def end_label(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return "present" if dt.date().isoformat() == TODAY else dt.date().isoformat()


def make_series_query(table: str, sensor_id: int, variable_id: int, start_date: str | None) -> str:
    where_lines = [
        f"    dv.sensorid = {sensor_id}",
        f"    AND dv.variableid = {variable_id}",
    ]
    if start_date:
        where_lines.append(f"    AND dv.localdatetime >= DATE '{start_date}'")
    return (
        "SELECT\n"
        "    dv.localdatetime,\n"
        "    dv.datavalue\n"
        f"FROM\n    {table} dv\n"
        "WHERE\n"
        + "\n".join(where_lines)
        + "\nORDER BY\n"
        "    dv.localdatetime;"
    )


def make_variable_query(schema_prefix: str, variable_id: int) -> str:
    return (
        "SELECT\n"
        "    v.variableid,\n"
        "    v.variablecode,\n"
        "    v.variablename,\n"
        "    u.unitsabbreviation\n"
        f"FROM\n    {schema_prefix}.variables v\n"
        f"    LEFT JOIN {schema_prefix}.units u\n"
        "        ON v.variableunitsid = u.unitsid\n"
        "WHERE\n"
        f"    v.variableid = {variable_id};"
    )


def fetch_oracle_sensor_info(cur: oracledb.Cursor, schema: str) -> dict[int, OracleSensorInfo]:
    cur.execute(
        f"""
        SELECT
            s.sensorid,
            s.sensorcode,
            s.datatablename,
            l.locationname,
            l.localx,
            l.localy,
            l.localz,
            l.boxz,
            l.boxverticaldatum,
            l.dlevel,
            l.elevation
        FROM {schema}.sensors s
            LEFT JOIN {schema}.locations l
                ON s.locationid = l.locationid
        """
    )
    return {
        int(row[0]): OracleSensorInfo(
            sensor_code=row[1],
            data_table_name=row[2],
            location_name=row[3],
            local_x=row[4],
            local_y=row[5],
            local_z=row[6],
            box_z=row[7],
            box_vertical_datum=row[8],
            dlevel=row[9],
            elevation=row[10],
        )
        for row in cur.fetchall()
    }


def fetch_oracle_variable_info(cur: oracledb.Cursor, schema: str) -> dict[int, OracleVariableInfo]:
    cur.execute(
        f"""
        SELECT
            v.variableid,
            v.variablecode,
            v.variablename,
            u.unitsabbreviation
        FROM {schema}.variables v
            LEFT JOIN {schema}.units u
                ON v.variableunitsid = u.unitsid
        """
    )
    return {
        int(row[0]): OracleVariableInfo(
            variable_id=int(row[0]),
            variable_code=row[1],
            variable_name=row[2],
            units=row[3],
        )
        for row in cur.fetchall()
    }


def fetch_catalog_bounds(
    cur: oracledb.Cursor,
    schema: str,
    variable_ids: list[int],
) -> dict[tuple[int, int], CatalogBounds]:
    if not variable_ids:
        return {}
    placeholders = ", ".join(str(int(variable_id)) for variable_id in sorted(set(variable_ids)))
    cur.execute(
        f"""
        SELECT
            sensorid,
            variableid,
            MIN(dateday),
            MAX(dateday),
            SUM(valuecnt)
        FROM {schema}.datacatalog
        WHERE variableid IN ({placeholders})
        GROUP BY sensorid, variableid
        """
    )
    return {
        (int(row[0]), int(row[1])): CatalogBounds(
            first_day=str(row[2]),
            last_day=str(row[3]),
            value_count=int(row[4] or 0),
        )
        for row in cur.fetchall()
    }


def parse_catalog_day(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def fetch_bounds_sensor(
    cur: oracledb.Cursor,
    table: str,
    sensor_id: int,
    variable_id: int,
    start_date: str | None,
    catalog_bounds: CatalogBounds | None = None,
) -> SeriesBounds:
    if table.endswith(".datavalues") and catalog_bounds is not None:
        return SeriesBounds(
            first_dt=parse_catalog_day(catalog_bounds.first_day),
            first_val=None,
            last_dt=parse_catalog_day(catalog_bounds.last_day),
            last_val=None,
            validation_source="Oracle DATACATALOG",
        )

    start_filter = f"\n          AND dv.localdatetime >= DATE '{start_date}'" if start_date else ""
    q_first = f"""
        SELECT dv.localdatetime, dv.datavalue
        FROM {table} dv
        WHERE dv.sensorid = :sensor_id
          AND dv.variableid = :variable_id
          {start_filter}
        ORDER BY dv.localdatetime
        FETCH FIRST 1 ROW ONLY
    """
    q_last = f"""
        SELECT dv.localdatetime, dv.datavalue
        FROM {table} dv
        WHERE dv.sensorid = :sensor_id
          AND dv.variableid = :variable_id
          {start_filter}
        ORDER BY dv.localdatetime DESC
        FETCH FIRST 1 ROW ONLY
    """
    try:
        cur.execute(q_first, sensor_id=sensor_id, variable_id=variable_id)
        first_row = cur.fetchone()
        cur.execute(q_last, sensor_id=sensor_id, variable_id=variable_id)
        last_row = cur.fetchone()
    except oracledb.Error as exc:
        return SeriesBounds(
            first_dt=None,
            first_val=None,
            last_dt=None,
            last_val=None,
            validation_error=f"{type(exc).__name__}: {exc}",
        )
    return SeriesBounds(
        first_dt=first_row[0] if first_row else None,
        first_val=first_row[1] if first_row else None,
        last_dt=last_row[0] if last_row else None,
        last_val=last_row[1] if last_row else None,
    )


def classify_row(variable_name: str, variable_code: str, sheet_name: str) -> dict[str, str]:
    if variable_name == "Relative humidity":
        return {
            "quantity": "Relative humidity",
            "symbol": "RH",
            "role": "Atmospheric moisture boundary driver",
            "how": "Air relative humidity constrains atmospheric moisture state and gas-transfer boundary conditions.",
            "why": "Useful for interpreting evaporation, near-surface gas exchange, and moisture controls on CO2 transport.",
        }
    if variable_name == "Volumetric water content":
        return {
            "quantity": "Volumetric water content",
            "symbol": "theta",
            "role": "Moisture state / transport driver",
            "how": "Volumetric water content controls gas-filled porosity and effective diffusivity in the porous medium.",
            "why": "Core moisture measurement for coupling hydrologic state to subsurface gas transport.",
        }
    if variable_name == "Water potential":
        return {
            "quantity": "Water potential",
            "symbol": "psi_m",
            "role": "Hydrologic state support",
            "how": "Water potential describes matric state and helps interpret water retention and drainage controls.",
            "why": "Useful support variable for explaining moisture redistribution and pore-space connectivity.",
        }
    if variable_name == "Water vapor concentration":
        return {
            "quantity": "Water vapor concentration",
            "symbol": "C_H2O",
            "role": "Atmospheric moisture support",
            "how": "Water vapor concentration records gas-phase moisture in the same LI-COR air stream.",
            "why": "Useful for interpreting humidity effects and LI-COR gas measurements near atmospheric CO2 rows.",
        }
    if variable_name == "Temperature" and variable_code == "AirTemp":
        return {
            "quantity": "Air temperature",
            "symbol": "T_air",
            "role": "Atmospheric thermal boundary driver",
            "how": "Air temperature constrains atmospheric boundary state and temperature-dependent gas properties.",
            "why": "Core atmospheric thermal measurement for interpreting surface exchange and sensor response.",
        }
    if variable_name == "Temperature" and variable_code == "waterTemp":
        return {
            "quantity": "Water temperature",
            "symbol": "T_water",
            "role": "Hydrologic thermal support",
            "how": "Water temperature supports interpretation of subsurface water-level and drainage measurements.",
            "why": "Useful for hydrologic context and temperature-dependent water/gas transport properties.",
        }
    if variable_name == "Temperature" and variable_code in {"DeviceTemp", "IRGA_Temp", "PyrgeoTemp"}:
        return {
            "quantity": "Instrument temperature",
            "symbol": "T_instrument",
            "role": "Instrument diagnostics / support",
            "how": "Instrument temperature tracks measurement-environment or electronics thermal state.",
            "why": "Diagnostic support for interpreting temperature-sensitive sensors and quality-control issues.",
        }
    if variable_name == "Temperature" and variable_code in {"SurfaceTemp", "SonicTemp", "SonicTempAvg"}:
        return {
            "quantity": "Surface / sonic temperature",
            "symbol": "T",
            "role": "Thermal boundary support",
            "how": "Surface or sonic temperature helps characterize thermal forcing near the slope/air interface.",
            "why": "Useful support for interpreting thermal controls on gas transport and atmospheric coupling.",
        }
    if variable_name == "Temperature" and sheet_name in {"5TM", "MPS-2", "TCAV"}:
        return {
            "quantity": "Soil temperature",
            "symbol": "T_soil",
            "role": "Thermal state / transport driver",
            "how": "Subsurface temperature controls gas diffusivity, density, and biological or geochemical source-sink rates.",
            "why": "Core thermal state measurement for coupling heat and moisture to subsurface gas transport.",
        }
    return {
        "quantity": variable_name,
        "symbol": "T" if variable_name == "Temperature" else variable_code,
        "role": "Thermal / moisture support",
        "how": "Temperature or moisture time series used as a support driver for transport interpretation.",
        "why": "Useful for interpreting environmental controls on CO2 transport and sensor behavior.",
    }


def spatial_fields(
    sheet_name: str,
    level: int | None,
    x_coord: float | None,
    y_coord: float | None,
    info: OracleSensorInfo | None,
) -> tuple[float | None, str, str]:
    if sheet_name in {"HMP60", "LI-COR"}:
        if level in AIR_LEVEL_HEIGHT_M:
            height_m = AIR_LEVEL_HEIGHT_M[level]
            return (
                height_m,
                f"Level {level} ({height_m * 100:g} cm above surface)",
                f"Atmospheric point above the slope at x={x_coord}, y={y_coord}.",
            )
        return (None, f"Level {level}" if level is not None else "", "Atmospheric system/control point.")

    datum = (info.box_vertical_datum if info else "") or ""
    box_z = as_float(info.box_z if info else None)
    if "depth" in datum.casefold() and box_z is not None:
        depth_cm = box_z * 100
        z_coord = -box_z
        if level and level > 0:
            label = f"Depth level {level} ({depth_cm:g} cm below surface)"
        else:
            label = f"{depth_cm:g} cm below surface"
        return (z_coord, label, f"Subsurface point at x={x_coord}, y={y_coord}, depth={depth_cm:g} cm.")
    if "height" in datum.casefold() and box_z is not None:
        height_cm = box_z * 100
        return (box_z, f"Level {level} ({height_cm:g} cm above surface)", f"Air/surface point at x={x_coord}, y={y_coord}.")
    if sheet_name in {"CNR4", "Control", "ShortTermSet", "CSAT3"}:
        return (0, "surface / instrument level", f"Surface or instrument-support point at x={x_coord}, y={y_coord}.")
    return (None, f"Level {level}" if level is not None else "", f"Point at x={x_coord}, y={y_coord}.")


def should_keep(variable_name: str | None) -> bool:
    return str(variable_name or "").strip() in KEEP_VARIABLE_NAMES


def table_name(schema_prefix: str, sheet_name: str, info: OracleSensorInfo | None) -> str:
    raw_name = (info.data_table_name if info else None) or ""
    if not raw_name:
        raw_name = "datavalueslicor" if sheet_name == "LI-COR" else "datavalues"
    return f"{schema_prefix}.{raw_name.strip().lower()}"


def iter_inventory_candidates(slope: dict[str, Any]) -> list[dict[str, Any]]:
    wb = load_workbook(slope["inventory"], read_only=True, data_only=True)
    candidates: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if ws.title == "SensorTypes":
            continue
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        indexes = header_indexes(header)
        idx_sensor_id = get_index(indexes, "SENSORID")
        idx_sensor_code = get_index(indexes, "SENSORCODE")
        idx_sensor_type = get_index(indexes, "SENSORTYPE")
        idx_variable_id = get_index(indexes, "VARIABLEID")
        idx_variable_code = get_index(indexes, "VARIABLECODE")
        idx_variable_name = get_index(indexes, "VARIABLENAME")
        idx_units = get_index(indexes, "UNITS", "UNITSABBREVIATION")
        idx_data_from = get_index(indexes, "Data From", last=True)
        idx_data_end = get_index(indexes, "Data End", last=True)
        idx_x = get_index(indexes, "X-across")
        idx_y = get_index(indexes, "Y-upslope")
        idx_level = get_index(indexes, "Depth level", "Level")
        if idx_sensor_id is None or idx_variable_id is None or idx_variable_name is None:
            continue
        for row in rows:
            sensor_id = as_int(row_value(row, idx_sensor_id))
            variable_id = as_int(row_value(row, idx_variable_id))
            variable_name = str(row_value(row, idx_variable_name) or "").strip()
            if sensor_id is None or variable_id is None or not should_keep(variable_name):
                continue
            candidates.append(
                {
                    "slope_label": slope["label"],
                    "schema": slope["schema"],
                    "schema_prefix": slope["schema_prefix"],
                    "inventory_file": slope["inventory"].name,
                    "sheet_name": ws.title,
                    "sensor_id": sensor_id,
                    "sensor_code": row_value(row, idx_sensor_code),
                    "sensor_type": row_value(row, idx_sensor_type),
                    "variable_id": variable_id,
                    "variable_code": str(row_value(row, idx_variable_code) or "").strip(),
                    "variable_name": variable_name,
                    "units": row_value(row, idx_units),
                    "inventory_start": row_value(row, idx_data_from),
                    "inventory_end": row_value(row, idx_data_end),
                    "x_coord": as_float(row_value(row, idx_x)),
                    "y_coord": as_float(row_value(row, idx_y)),
                    "level": as_int(row_value(row, idx_level)),
                }
            )
    return candidates


def build_record(
    cur: oracledb.Cursor,
    candidate: dict[str, Any],
    sensor_info: dict[int, OracleSensorInfo],
    variable_info: dict[int, OracleVariableInfo],
    catalog_info: dict[tuple[int, int], CatalogBounds],
) -> dict[str, Any]:
    sensor_id = int(candidate["sensor_id"])
    variable_id = int(candidate["variable_id"])
    info = sensor_info.get(sensor_id)
    variable = variable_info.get(variable_id)
    variable_code = variable.variable_code if variable and variable.variable_code else candidate["variable_code"]
    variable_name = variable.variable_name if variable and variable.variable_name else candidate["variable_name"]
    units = variable.units if variable and variable.units else candidate["units"]
    x_coord = candidate["x_coord"] if candidate["x_coord"] is not None else (info.local_x if info else None)
    y_coord = candidate["y_coord"] if candidate["y_coord"] is not None else (info.local_y if info else None)
    z_coord, height_depth, spatial_meaning = spatial_fields(
        candidate["sheet_name"],
        candidate["level"],
        x_coord,
        y_coord,
        info,
    )
    table = table_name(candidate["schema_prefix"], candidate["sheet_name"], info)
    inventory_start = date_label(candidate["inventory_start"])
    start_for_query = inventory_start if inventory_start and inventory_start != "present" else None
    catalog_bounds = catalog_info.get((sensor_id, variable_id))
    bounds = fetch_bounds_sensor(cur, table, sensor_id, variable_id, start_for_query, catalog_bounds)
    actual_start = date_label(bounds.first_dt)
    query_start = actual_start or start_for_query
    classification = classify_row(str(variable_name), str(variable_code), str(candidate["sheet_name"]))
    if bounds.has_data:
        availability_start = actual_start
        availability_end = end_label(bounds.last_dt)
        if bounds.validation_error:
            live_tested = (
                f"Partial - Oracle DATACATALOG reports rows, but raw latest-row validation had an error on {TODAY}."
            )
            note = (
                f"Oracle DATACATALOG reports data for this row, but raw latest-row validation failed on {TODAY}: "
                f"{bounds.validation_error}. Inventory window: {inventory_start or 'unknown'} to "
                f"{date_label(candidate['inventory_end']) or 'unknown'}."
            )
        else:
            live_tested = (
                f"Yes - {bounds.validation_source} validated in Oracle on {TODAY}."
            )
            note = (
                f"{bounds.validation_source} validated in Oracle on {TODAY}. "
                f"Inventory window: {inventory_start or 'unknown'} to {date_label(candidate['inventory_end']) or 'unknown'}."
            )
    else:
        availability_start = ""
        availability_end = ""
        live_tested = NO_ZERO_ROWS if not bounds.validation_error else f"No - Oracle validation error on {TODAY}."
        note = (
            f"Sensor metadata found in {candidate['inventory_file']} / {candidate['sheet_name']}, "
            f"but the Oracle validation returned zero rows or failed on {TODAY}. "
            f"{bounds.validation_error or ''} "
            f"Inventory window: {inventory_start or 'unknown'} to {date_label(candidate['inventory_end']) or 'unknown'}."
        )
    sensor_code = str(candidate["sensor_code"] or (info.sensor_code if info else "") or "").strip()
    location_name = info.location_name if info and info.location_name else candidate["slope_label"]
    return {
        "A": classification["quantity"],
        "B": classification["symbol"],
        "C": x_coord,
        "D": y_coord,
        "E": z_coord,
        "F": height_depth,
        "G": classification["role"],
        "H": classification["how"],
        "I": candidate["slope_label"],
        "J": candidate["inventory_file"],
        "K": candidate["sheet_name"],
        "L": sensor_id,
        "M": sensor_code,
        "N": table,
        "O": f"dv.sensorid = {sensor_id} AND dv.variableid = {variable_id}",
        "P": "LOCALDATETIME",
        "Q": "DATAVALUE",
        "R": units,
        "S": units,
        "T": "identity (keep physical units in SQL; convert later in Python)",
        "U": location_name,
        "V": spatial_meaning,
        "W": classification["why"],
        "X": availability_start,
        "Y": availability_end,
        "Z": note,
        "AA": make_series_query(table, sensor_id, variable_id, query_start),
        "AB": None,
        "AC": live_tested,
        "AD": variable_id,
        "AE": variable_code,
        "AF": variable_name,
        "AG": units,
        "AH": make_variable_query(candidate["schema_prefix"], variable_id),
        "AI": None,
        "AJ": None,
        "AK": None,
    }


def record_sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
    group_order = {
        "Soil temperature": 0,
        "Volumetric water content": 1,
        "Water potential": 2,
        "Air temperature": 3,
        "Relative humidity": 4,
        "Water vapor concentration": 5,
        "Water temperature": 6,
        "Surface / sonic temperature": 7,
        "Instrument temperature": 8,
    }
    slope_order = {"LEO East": 0, "LEO Center": 1, "LEO West": 2}
    return (
        slope_order.get(record["I"], 99),
        group_order.get(record["A"], 99),
        str(record["K"]),
        record["D"] if record["D"] is not None else 9999,
        record["C"] if record["C"] is not None else 9999,
        record["E"] if record["E"] is not None else 9999,
        record["L"],
        record["AD"],
    )


def copy_row_style(source_ws, target_ws, template_row: int, target_row: int) -> None:
    for col_idx in range(1, source_ws.max_column + 1):
        src = source_ws.cell(template_row, col_idx)
        dst = target_ws.cell(target_row, col_idx)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)
        dst.number_format = src.number_format
    target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[template_row].height


def copy_sheet_structure(wb, sheet_name: str, after_sheet: str) -> Any:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    source = wb["CO2"]
    if after_sheet in wb.sheetnames:
        insert_at = wb.sheetnames.index(after_sheet) + 1
    else:
        insert_at = len(wb.sheetnames)
    target = wb.create_sheet(sheet_name, insert_at)
    for merged_range in source.merged_cells.ranges:
        if merged_range.max_row <= 3:
            target.merge_cells(str(merged_range))
    for col_idx in range(1, source.max_column + 1):
        letter = get_column_letter(col_idx)
        target.column_dimensions[letter].width = source.column_dimensions[letter].width
    for row_idx in range(1, 4):
        target.row_dimensions[row_idx].height = source.row_dimensions[row_idx].height
        for col_idx in range(1, source.max_column + 1):
            src = source.cell(row_idx, col_idx)
            dst = target.cell(row_idx, col_idx)
            if isinstance(dst, MergedCell):
                continue
            dst.value = src.value
            dst._style = copy.copy(src._style)
            dst.font = copy.copy(src.font)
            dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.protection = copy.copy(src.protection)
            dst.number_format = src.number_format
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    return target


def write_records(records: list[dict[str, Any]], sheet_name: str, after_sheet: str) -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = copy_sheet_structure(wb, sheet_name, after_sheet)
    template = wb["CO2"]
    columns = [get_column_letter(i) for i in range(1, template.max_column + 1)]
    start_row = 4
    for row_index, record in enumerate(records, start=start_row):
        copy_row_style(template, ws, 4, row_index)
        for col in columns:
            ws[f"{col}{row_index}"] = record.get(col)
    last_row = max(3, start_row + len(records) - 1)
    ws.auto_filter.ref = f"A3:AK{last_row}"
    wb.save(WORKBOOK_PATH)


def main() -> None:
    conn = connect()
    cur = conn.cursor()
    records: list[dict[str, Any]] = []
    try:
        for slope in SLOPES:
            print(f"Reading {slope['inventory'].name}...", flush=True)
            candidates = iter_inventory_candidates(slope)
            print(f"  candidates: {len(candidates)}", flush=True)
            sensor_info = fetch_oracle_sensor_info(cur, slope["schema"])
            variable_info = fetch_oracle_variable_info(cur, slope["schema"])
            variable_ids = [int(candidate["variable_id"]) for candidate in candidates]
            catalog_info = fetch_catalog_bounds(cur, slope["schema"], variable_ids)
            print(f"  catalog bounds: {len(catalog_info)}", flush=True)
            for idx, candidate in enumerate(candidates, 1):
                if idx % 250 == 0:
                    print(
                        "  validating "
                        f"{idx}/{len(candidates)} "
                        f"{candidate['sheet_name']} "
                        f"sensorid={candidate['sensor_id']} "
                        f"variableid={candidate['variable_id']}",
                        flush=True,
                    )
                records.append(build_record(cur, candidate, sensor_info, variable_info, catalog_info))
        records.sort(key=record_sort_key)
        soil_temp_records = [record for record in records if record.get("AE") == "soilTemp"]
        humidity_records = [record for record in records if record.get("AE") in HUMIDITY_VARIABLE_CODES]
        moss_records = [
            record
            for record in records
            if record.get("AE") != "soilTemp" and record.get("AE") not in HUMIDITY_VARIABLE_CODES
        ]
        print(f"Writing {len(moss_records)} Moss rows to {WORKBOOK_PATH}...", flush=True)
        write_records(moss_records, "Moss", "CO2")
        print(f"Writing {len(humidity_records)} Humidity rows to {WORKBOOK_PATH}...", flush=True)
        write_records(humidity_records, "Humidity", "Moss")
        print(f"Writing {len(soil_temp_records)} soilTemp rows to {WORKBOOK_PATH}...", flush=True)
        write_records(soil_temp_records, "soilTemp", "Humidity")
    finally:
        cur.close()
        conn.close()
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
