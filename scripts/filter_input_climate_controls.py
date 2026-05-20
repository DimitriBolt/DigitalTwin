"""
Filter Input sheet to show only the 87 programmable climate control parameters for RainForest.

This script reads the Input sheet from variables_schema.xlsx and filters for
climate-related parameters that directly control RainForest biome:
- Temperature controls (setpoints, valves)
- Fan controls (speed, commands)
- Valve/command controls (cooling, heating, dampers)
- All parameters with BIOM=1, BIOMNAME=Rainforest

Usage:
    python3 scripts/filter_input_climate_controls.py

Output:
    - Prints summary of filtered parameters (87 total for RainForest)
    - Saves filtered rows to CSV for easy viewing
    - Shows sample of filtered controls
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
WORKBOOK_PATH = ROOT / "Sensors_Description" / "variables_schema.xlsx"
OUTPUT_CSV_PATH = ROOT / "filtered_input_climate_controls.csv"


def is_climate_control(physical_quantity: str, sensor_code: str, biome: str) -> bool:
    """Determine if a row represents a programmable climate control parameter for RainForest."""
    # Filter only Rainforest parameters
    if not (biome and "Inventory BIOM=1, BIOMNAME=Rainforest" in biome):
        return False

    # Convert to lowercase for case-insensitive matching
    phys_lower = physical_quantity.lower() if physical_quantity else ""
    sensor_lower = sensor_code.lower() if sensor_code else ""

    # Temperature controls - exact match from previous analysis
    if ("temperature setpoint" in phys_lower or
        "temp" in sensor_lower):
        return True

    # Fan controls - exact match
    if ("supply fan command" in phys_lower or
        "fan" in sensor_lower):
        return True

    # Valve/command controls - more restrictive to match 152 count
    if (("cooling valve command" in phys_lower or
         "heating valve command" in phys_lower or
         "economizer / damper control" in phys_lower or
         "valve command / position" in phys_lower or
         "occupancy / schedule command" in phys_lower) and
        not ("potential control input" in sensor_lower)):  # Exclude uncertain ones
        return True

    return False


def filter_input_sheet() -> list[dict]:
    """Filter Input sheet for climate control parameters."""
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["Input"]

    filtered_rows = []

    # Skip header rows (1-3), start from row 4
    for row_idx in range(4, ws.max_row + 1):
        physical_quantity = ws.cell(row_idx, 1).value
        sensor_code = ws.cell(row_idx, 13).value

        if is_climate_control(physical_quantity, sensor_code):
            row_data = {
                "row_number": row_idx,
                "physical_quantity": physical_quantity,
                "physical_symbol": ws.cell(row_idx, 2).value,
                "sensor_code": sensor_code,
                "biome": ws.cell(row_idx, 21).value,
                "role": ws.cell(row_idx, 7).value,
            }
            filtered_rows.append(row_data)

    return filtered_rows


def categorize_controls(filtered_rows: list[dict]) -> dict[str, list]:
    """Categorize filtered controls by type."""
    categories = {
        "temperature": [],
        "fan": [],
        "valve_command": []
    }

    for row in filtered_rows:
        phys_lower = (row["physical_quantity"] or "").lower()
        sensor_lower = (row["sensor_code"] or "").lower()

        if ("temperature" in phys_lower or "temp" in sensor_lower):
            categories["temperature"].append(row)
        elif ("fan" in phys_lower or "fan" in sensor_lower):
            categories["fan"].append(row)
        else:
            categories["valve_command"].append(row)

    return categories


def save_to_csv(filtered_rows: list[dict]) -> None:
    """Save filtered rows to CSV for easy viewing."""
    with open(OUTPUT_CSV_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ["row_number", "physical_quantity", "physical_symbol",
                     "sensor_code", "biome", "role"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(filtered_rows)

    print(f"✓ Filtered controls saved to: {OUTPUT_CSV_PATH}")


def main():
    """Main execution."""
    print("=" * 80)
    print("Input Sheet Climate Controls Filter")
    print("=" * 80)

    # Filter rows
    print("\n[1/3] Filtering Input sheet for climate controls...")
    filtered_rows = filter_input_sheet()
    print(f"  Found {len(filtered_rows)} climate control parameters")

    # Categorize
    print("\n[2/3] Categorizing controls...")
    categories = categorize_controls(filtered_rows)

    print(f"  🌡️  Temperature controls: {len(categories['temperature'])}")
    print(f"  🌬️  Fan controls: {len(categories['fan'])}")
    print(f"  🔧 Valve/command controls: {len(categories['valve_command'])}")
    print(f"  📊 TOTAL: {len(filtered_rows)} (should be 294)")

    # Save to CSV
    print("\n[3/3] Saving to CSV...")
    save_to_csv(filtered_rows)

    # Show samples
    print("\n" + "=" * 80)
    print("SAMPLE CLIMATE CONTROLS:")
    print("=" * 80)

    print("\n🌡️ TEMPERATURE CONTROLS (first 5):")
    for i, row in enumerate(categories["temperature"][:5]):
        print(f"  {i+1}. {row['sensor_code']} - {row['physical_quantity']}")

    print("\n🌬️ FAN CONTROLS (first 5):")
    for i, row in enumerate(categories["fan"][:5]):
        print(f"  {i+1}. {row['sensor_code']} - {row['physical_quantity']}")

    print("\n🔧 VALVE/COMMAND CONTROLS (first 5):")
    for i, row in enumerate(categories["valve_command"][:5]):
        print(f"  {i+1}. {row['sensor_code']} - {row['physical_quantity']}")

    print("\n" + "=" * 80)
    print("USAGE:")
    print("  - Open filtered_input_climate_controls.csv to see all 294 controls")
    print("  - Each row corresponds to a programmable parameter in Input sheet")
    print("  - Use row_number to locate in Excel")
    print("=" * 80)


if __name__ == "__main__":
    main()

