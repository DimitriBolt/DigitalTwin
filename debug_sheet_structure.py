"""
Debug script to check actual structure of Input and Output sheets.
"""

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
WORKBOOK_PATH = ROOT / "Sensors_Description" / "variables_schema.xlsx"

wb = load_workbook(WORKBOOK_PATH, data_only=False)

# Check Input sheet
print("=" * 80)
print("INPUT SHEET - First 10 rows (starting from row 4)")
print("=" * 80)
ws = wb["Input"]
for i, row_idx in enumerate(range(4, 14)):
    phys_qty = ws.cell(row_idx, 1).value
    role = ws.cell(row_idx, 7).value
    sensor_code = ws.cell(row_idx, 13).value

    # Check columns around U
    biome_21 = ws.cell(row_idx, 21).value
    biome_22 = ws.cell(row_idx, 22).value
    biome_23 = ws.cell(row_idx, 23).value

    print(f"\nRow {row_idx}:")
    print(f"  A: {phys_qty}")
    print(f"  G: {role}")
    print(f"  M: {sensor_code}")
    print(f"  Col 21 (U): {biome_21}")
    print(f"  Col 22 (V): {biome_22}")
    print(f"  Col 23 (W): {biome_23}")

# Find which column has BIOMNAME
print("\n" + "=" * 80)
print("Searching for BIOMNAME column in Input sheet...")
print("=" * 80)
for col_idx in range(1, 40):
    header_val = ws.cell(3, col_idx).value
    if header_val and "BIOM" in str(header_val):
        print(f"Column {col_idx}: {header_val}")
        # Show first value from this column
        first_val = ws.cell(4, col_idx).value
        print(f"  First value (row 4): {first_val}")

# Check Output sheet
print("\n" + "=" * 80)
print("OUTPUT SHEET - First 10 rows (starting from row 4)")
print("=" * 80)
ws = wb["Output"]
for i, row_idx in enumerate(range(4, 14)):
    phys_qty = ws.cell(row_idx, 1).value
    role = ws.cell(row_idx, 7).value
    sensor_code = ws.cell(row_idx, 13).value

    biome_21 = ws.cell(row_idx, 21).value
    biome_22 = ws.cell(row_idx, 22).value
    biome_23 = ws.cell(row_idx, 23).value

    print(f"\nRow {row_idx}:")
    print(f"  A: {phys_qty}")
    print(f"  G: {role}")
    print(f"  M: {sensor_code}")
    print(f"  Col 21 (U): {biome_21}")
    print(f"  Col 22 (V): {biome_22}")
    print(f"  Col 23 (W): {biome_23}")

# Find which column has BIOMNAME in Output
print("\n" + "=" * 80)
print("Searching for BIOMNAME column in Output sheet...")
print("=" * 80)
for col_idx in range(1, 40):
    header_val = ws.cell(3, col_idx).value
    if header_val and "BIOM" in str(header_val):
        print(f"Column {col_idx}: {header_val}")
        # Show first value from this column
        first_val = ws.cell(4, col_idx).value
        print(f"  First value (row 4): {first_val}")

